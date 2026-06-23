import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "店长减负调查问卷"

# 样式定义
header_font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
section_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
question_font = Font(name="微软雅黑", size=11, bold=True)
option_font = Font(name="微软雅黑", size=10)
note_font = Font(name="微软雅黑", size=9, italic=True, color="666666")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

current_row = 1

def write_header(text, row, colspan=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
    for c in range(2, colspan+1):
        ws.cell(row=row, column=c).border = thin_border
    return row + 1

def write_section(text, row, colspan=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border
    for c in range(2, colspan+1):
        ws.cell(row=row, column=c).border = thin_border
    return row + 1

def write_question(text, row, colspan=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = question_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = thin_border
    for c in range(2, colspan+1):
        ws.cell(row=row, column=c).border = thin_border
    return row + 1

def write_options(options, row, start_col=1, is_matrix=False, headers=None):
    if is_matrix and headers:
        # 矩阵题表头
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=start_col+i, value=h)
            cell.font = option_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        row += 1
        for opt in options:
            for i, val in enumerate(opt):
                cell = ws.cell(row=row, column=start_col+i, value=val)
                cell.font = option_font
                cell.alignment = Alignment(horizontal="left" if i==0 else "center", vertical="center", wrap_text=True)
                cell.border = thin_border
            row += 1
    else:
        for opt in options:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=10)
            cell = ws.cell(row=row, column=start_col, value=opt)
            cell.font = option_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
            for c in range(2, 11):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
    return row

def write_note(text, row, colspan=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = note_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    return row + 1

def write_open(text, row, colspan=10, lines=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row+lines-1, end_column=colspan)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = option_font
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell.border = thin_border
    for r in range(row, row+lines):
        for c in range(2, colspan+1):
            ws.cell(row=r, column=c).border = thin_border
    return row + lines

# ==================== 标题 ====================
current_row = write_header("店长门店管理与带训效率调查问卷", current_row)
current_row = write_note("调查目的：识别店长日常工作中的繁琐、低效环节，以及带训徒弟过程中的痛点和工具需求，为'店长减负、提升带训效率'项目提供数据支撑。", current_row)
current_row = write_note("预计填写时间：8-12分钟    |    填写对象：各门店店长/副店长/领班/储备店长", current_row)
current_row += 1

# ==================== 第一部分：基本信息 ====================
current_row = write_section("第一部分：基本信息", current_row)

current_row = write_question("1. 您的门店：", current_row)
current_row = write_open("请填写门店名称", current_row, lines=1)

current_row = write_question("2. 您的岗位：", current_row)
current_row = write_options(["□ 店长", "□ 副店长/领班", "□ 储备店长"], current_row)

current_row = write_question("3. 带训徒弟年限：", current_row)
current_row = write_options(["□ 无带训经验", "□ 1年以内", "□ 1-3年", "□ 3年以上"], current_row)

current_row = write_question("4. 当前门店员工数（含兼职）：", current_row)
current_row = write_open("____人", current_row, lines=1)
current_row += 1

# ==================== 第二部分：日常工作耗时分布 ====================
current_row = write_section("第二部分：日常工作耗时分布", current_row)
current_row = write_note("请估算您平均每周在以下事项上花费的时间占比", current_row)

matrix_headers = ["工作类别", "几乎不占(0-5%)", "较少(5-15%)", "中等(15-30%)", "较多(30-50%)", "非常耗时(50%+)", "", "", "", ""]
matrix_options = [
    ["1. 现场服务与出品管控", "○", "○", "○", "○", "○", "", "", "", ""],
    ["2. 人员排班、考勤、请假管理", "○", "○", "○", "○", "○", "", "", "", ""],
    ["3. 业绩数据记录、统计、分析", "○", "○", "○", "○", "○", "", "", "", ""],
    ["4. 报销单、费用单据整理", "○", "○", "○", "○", "○", "", "", "", ""],
    ["5. 各类检查表、报表填写", "○", "○", "○", "○", "○", "", "", "", ""],
    ["6. 群内消息回复、沟通协调", "○", "○", "○", "○", "○", "", "", "", ""],
    ["7. 培训带教、新人指导", "○", "○", "○", "○", "○", "", "", "", ""],
    ["8. 产品报货、损耗记录", "○", "○", "○", "○", "○", "", "", "", ""],
    ["9. 其他行政事务", "○", "○", "○", "○", "○", "", "", "", ""],
]
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
cell = ws.cell(row=current_row, column=1, value="")
cell.border = thin_border
for c in range(2, 11):
    ws.cell(row=current_row, column=c).border = thin_border
current_row += 1

current_row = write_options(matrix_options, current_row, is_matrix=True, headers=matrix_headers)

current_row = write_question("追问：以上哪3项最让您感到'繁琐、耗时、想摆脱'？", current_row)
current_row = write_open("1. ______________  2. ______________  3. ______________", current_row, lines=2)
current_row += 1

# ==================== 第三部分：门店管理痛点 ====================
current_row = write_section("第三部分：门店管理痛点", current_row)

current_row = write_question("5. 以下哪些工作您目前需要'手工操作'且觉得特别麻烦？（可多选）", current_row)
current_row = write_options([
    "□ 手工统计每日营业额/来客数/客单价",
    "□ 手工填写日报/周报/月报",
    "□ 手工整理报销单、拍照、分类、贴票",
    "□ 手工核对考勤、计算工时",
    "□ 手工填写各类检查表（卫生/设备/巡店）",
    "□ 手工制作排班表",
    "□ 手工统计产品损耗/报货",
    "□ 其他：________________________"
], current_row)

current_row = write_question("6. 数据统计方面，您希望系统自动帮您做哪些事？（可多选）", current_row)
current_row = write_options([
    "□ 自动抓取营业额、来客数、客单价数据",
    "□ 自动对比目标/环比/同比",
    "□ 自动分析品类占比、时段占比、渠道占比",
    "□ 自动标红异常数据（如跌破目标的品类）",
    "□ 自动生成文字分析建议",
    "□ 其他：________________________"
], current_row)

current_row = write_question("7. 沟通协作方面，您遇到的最大困扰是？（最多选3项）", current_row)
current_row = write_options([
    "□ 钉钉群太多，消息爬楼困难，重要信息被淹没",
    "□ 跨部门找人难（工程/采购/人事/财务）",
    "□ 总部通知层层转达，容易遗漏或变形",
    "□ 客诉处理需要层层上报，流程慢",
    "□ 问一个问题要在多个群里重复发",
    "□ 其他：________________________"
], current_row)

current_row = write_question("8. 信息获取方面，您目前最大的问题是？（可多选）", current_row)
current_row = write_options([
    "□ 产品SOP、标准图不知道在哪查",
    "□ 新品信息、活动方案更新不及时",
    "□ 想看同期/同区域/同类店的数据对比，找不到",
    "□ 公司制度、人事流程、财务规范分散在不同地方",
    "□ 想学习优秀门店的经验，但不知道怎么获取",
    "□ 其他：________________________"
], current_row)
current_row += 1

# ==================== 第四部分：带训徒弟痛点 ====================
current_row = write_section("第四部分：带训徒弟痛点", current_row)

current_row = write_question("9. 您目前带训徒弟，通常怎么做？（可多选）", current_row)
current_row = write_options([
    "□ 口头讲解+现场示范",
    "□ 让徒弟跟着看，边做边学",
    "□ 给徒弟发纸质/电子版SOP文件",
    "□ 定期检查徒弟操作，指出问题",
    "□ 安排徒弟看视频学习",
    "□ 其他：________________________"
], current_row)

current_row = write_question("10. 带训过程中，最让您头疼的是？（最多选3项）", current_row)
current_row = write_options([
    "□ 没有统一、标准的带训流程，每个人教法不一样",
    "□ SOP文件太旧，和实际操作不符",
    "□ 徒弟基础差，简单动作要反复教很多次",
    "□ 自己忙的时候，没时间专门带徒弟",
    "□ 不知道徒弟学会了没有，没有检验标准",
    "□ 带训内容零散，没有阶段性目标和检查清单",
    "□ 徒弟犯错时，不知道按什么标准评估和纠正",
    "□ 其他：________________________"
], current_row)

current_row = write_question("11. 如果有一站式带训工具，您希望它能提供什么？（可多选）", current_row)
current_row = write_options([
    "□ 每个岗位的标准化学习清单（第1天学什么、第2天学什么...）",
    "□ 关键操作的视频示范（如烤炉调温、裱花手法）",
    "□ 每日带训记录模板，徒弟和师傅签字确认",
    "□ 阶段考核表，明确徒弟什么阶段必须掌握什么技能",
    "□ 常见问题库（徒弟常犯的错，师傅怎么纠正）",
    "□ 师傅带训技巧指南（怎么教、怎么反馈、怎么激励）",
    "□ 其他：________________________"
], current_row)

current_row = write_question("12. 您是否遇到过以下带训场景？觉得怎么解决更好？", current_row)
current_row = write_open("场景A：徒弟学了后面忘了前面，同一个问题反复犯。\n您现在的做法：________________________________\n希望得到的支持：________________________________", current_row, lines=3)
current_row = write_open("场景B：店里忙的时候，徒弟站在一边不知道干什么，师傅也没空教。\n您现在的做法：________________________________\n希望得到的支持：________________________________", current_row, lines=3)
current_row = write_open("场景C：徒弟出师后，操作还是不够标准。\n您现在的做法：________________________________\n希望得到的支持：________________________________", current_row, lines=3)
current_row += 1

# ==================== 第五部分：工具与赋能需求 ====================
current_row = write_section("第五部分：工具与赋能需求", current_row)

current_row = write_question("13. 如果让您选3个最希望AI/系统帮您解决的问题，您选？（最多选3项）", current_row)
current_row = write_options([
    "□ 自动整理报销单、生成报销文档",
    "□ 自动分析营业额数据，给出诊断建议",
    "□ 自动生成日报/周报，店长只需确认",
    "□ 一键搜索：想查什么制度/SOP，直接问AI就给出答案",
    "□ 自动提醒：今天该做什么检查、该带徒弟学什么",
    "□ 语音输入：我用嘴说，AI帮我整理成文字记录",
    "□ 业绩预警：营业额/来客数异常时自动发通知",
    "□ 带训助手：徒弟操作不规范，AI自动提示纠正方法",
    "□ 其他：________________________"
], current_row)

current_row = write_question("14. 您对'AI智能助教'进入门店，最大的顾虑是？（可多选）", current_row)
current_row = write_options([
    "□ 担心操作复杂，不会用",
    "□ 担心取代人工，影响岗位稳定性",
    "□ 担心数据泄露/隐私问题",
    "□ 担心AI建议不靠谱，误导决策",
    "□ 没什么顾虑，能用上最好",
    "□ 其他：________________________"
], current_row)
current_row += 1

# ==================== 第六部分：开放建议 ====================
current_row = write_section("第六部分：开放建议", current_row)

current_row = write_question("15. 如果让您给总部提一个'最迫切想解决的管理问题'，您会提什么？", current_row)
current_row = write_open("请在此填写：", current_row, lines=3)

current_row = write_question("16. 如果让您给总部提一个'最迫切想解决的带训问题'，您会提什么？", current_row)
current_row = write_open("请在此填写：", current_row, lines=3)
current_row += 1

# ==================== 结尾 ====================
current_row = write_header("感谢您的时间！您的反馈将直接用于设计'店长减负与带训提效'方案。", current_row)

# 设置列宽
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 12
ws.column_dimensions["G"].width = 10
ws.column_dimensions["H"].width = 10
ws.column_dimensions["I"].width = 10
ws.column_dimensions["J"].width = 10

# 设置行高（标题行）
for row in range(1, current_row+1):
    ws.row_dimensions[row].height = 28

# 保存
output_path = "/root/.openclaw/workspace/店长减负调查问卷.xlsx"
wb.save(output_path)
print(f"✅ 问卷已生成：{output_path}")
print(f"共 {current_row} 行，包含6个部分，16道题目")
